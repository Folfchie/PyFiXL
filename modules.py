import openpyxl as xl
import os
import shutil
from tkinter import *
from tkinter import ttk


def process_income_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        gross_income_cell = sheet.cell(row, column=2)
        tax_owed_cell = sheet.cell(row, 3)
        net_income_cell = sheet.cell(row, 4)
        if gross_income_cell.value or net_income_cell.value or tax_owed_cell.value is not None:
            tax_owed = gross_income_cell.value - net_income_cell.value
            tax_owed_cell.value = tax_owed

    wb.save(filename)
    print(f"\nThe operation on {filename} has finished!")


def process_mtd_figures(income_wb_name,
                        expenses_wb_name,
                        mtd_wb_name):
    income_wb = xl.load_workbook(income_wb_name)
    expenses_wb = xl.load_workbook(expenses_wb_name)
    mtd_wb = xl.load_workbook(mtd_wb_name)
    income_sheet, expenses_sheet, mtd_sheet = (income_wb['Sheet1'],
                                               expenses_wb['Sheet1'],
                                               mtd_wb['Sheet1'])
    total_gross_pay = 0
    total_tax_paid = 0
    total_net_pay = 0
    total_expenses = 0

    for row in range(2, income_sheet.max_row + 1):
        gross_cell = income_sheet.cell(row, column=2)
        tax_cell = income_sheet.cell(row, 3)
        net_cell = income_sheet.cell(row, 4)
        if gross_cell.value is None:
            total_gross_pay += 0
        else:
            total_gross_pay += gross_cell.value
        if tax_cell.value is None:
            total_tax_paid += 0
        else:
            total_tax_paid += tax_cell.value
        if net_cell.value is None:
            total_net_pay += 0
        else:
            total_net_pay += net_cell.value

    for row in range(2, expenses_sheet.max_row + 1):
        cost_cell = expenses_sheet.cell(row, 3)
        total_expenses += cost_cell.value

    total_savings = total_net_pay - total_expenses

    mtd_gross_cell = mtd_sheet.cell(row=2, column=1)
    mtd_tax_cell = mtd_sheet.cell(2, 2)
    mtd_net_cell = mtd_sheet.cell(2, 3)
    mtd_expenses_cell = mtd_sheet.cell(2, 4)
    mtd_savings_cell = mtd_sheet.cell(2, 5)
    mtd_gross_cell.value = total_gross_pay
    mtd_tax_cell.value = total_tax_paid
    mtd_net_cell.value = total_net_pay
    mtd_expenses_cell.value = total_expenses
    mtd_savings_cell.value = total_savings

    mtd_wb.save(mtd_wb_name)
    print(f"\nThe operation on {mtd_wb_name} has finished!")


def process_ytd_figures(mtd_dir, ytd_dir, year):
    start_dir = os.getcwd()
    mtd_names = os.listdir(mtd_dir)
    ytd_names = os.listdir(ytd_dir)
    ytd_wb_name = None
    os.chdir(ytd_dir)

    for filename in ytd_names:
        if 'ytd' and year in filename:
            ytd_wb_name = filename

    shutil.copy(src=ytd_wb_name, dst=mtd_dir)
    os.remove(ytd_wb_name)
    os.chdir(mtd_dir)
    ytd_wb = xl.load_workbook(ytd_wb_name)
    ytd_sheet = ytd_wb['Sheet1']

    for filename in mtd_names:
        if 'mtd' in filename:
            current_mtd_wb = xl.load_workbook(filename)
            current_mtd_sheet = current_mtd_wb['Sheet1']
            current_mtd_gross_income_cell = current_mtd_sheet.cell(row=2, column=1)
            current_mtd_tax_owed_cell = current_mtd_sheet.cell(row=2, column=2)
            current_mtd_net_income_cell = current_mtd_sheet.cell(row=2, column=3)
            current_mtd_expenses_cell = current_mtd_sheet.cell(row=2, column=4)
            current_mtd_savings_cell = current_mtd_sheet.cell(row=2, column=5)
            months = ['jan', 'feb', 'mar', 'apr', 'may', 'jun',
                      'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
            if months[0] in filename:
                current_ytd_row = 2
            elif months[1] in filename:
                current_ytd_row = 3
            elif months[2] in filename:
                current_ytd_row = 4
            elif months[3] in filename:
                current_ytd_row = 5
            elif months[4] in filename:
                current_ytd_row = 6
            elif months[5] in filename:
                current_ytd_row = 7
            elif months[6] in filename:
                current_ytd_row = 8
            elif months[7] in filename:
                current_ytd_row = 9
            elif months[8] in filename:
                current_ytd_row = 10
            elif months[9] in filename:
                current_ytd_row = 11
            elif months[10] in filename:
                current_ytd_row = 12
            elif months[11] in filename:
                current_ytd_row = 13
            else:
                current_ytd_row = None
            if current_ytd_row is not None:
                # Get cells and store in local variables
                current_ytd_gross_income_cell = ytd_sheet.cell(current_ytd_row, column=2)
                current_ytd_tax_owed_cell = ytd_sheet.cell(current_ytd_row, 3)
                current_ytd_net_income_cell = ytd_sheet.cell(current_ytd_row, 4)
                current_ytd_expenses_cell = ytd_sheet.cell(current_ytd_row, 5)
                current_ytd_savings_cell = ytd_sheet.cell(current_ytd_row, 6)
                # Copy values from mtd to ytd
                current_ytd_gross_income_cell.value = current_mtd_gross_income_cell.value
                current_ytd_tax_owed_cell.value = current_mtd_tax_owed_cell.value
                current_ytd_net_income_cell.value = current_mtd_net_income_cell.value
                current_ytd_expenses_cell.value = current_mtd_expenses_cell.value
                current_ytd_savings_cell.value = current_mtd_savings_cell.value

    ytd_wb.save(ytd_wb_name)
    shutil.copy(src=ytd_wb_name, dst=ytd_dir)
    os.remove(ytd_wb_name)
    os.chdir(start_dir)
    print(f"\nThe operation on {ytd_wb_name} has finished!")

def open_test_window():
    root = Tk()
    frm = ttk.Frame(root, padding=10)
    frm.grid()
    ttk.Label(frm, text="Hello World!").grid(column=0, row=0)
    ttk.Button(frm, text="Quit", command=root.destroy).grid(column=1, row=0)
    root.mainloop()